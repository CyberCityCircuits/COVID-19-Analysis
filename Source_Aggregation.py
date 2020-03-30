#Data Source: https://github.com/CSSEGISandData/COVID-19
#COVID-19 Aggregation Script by Cyber City Circuits

import csv, os
import openpyxl
from openpyxl import Workbook
import datetime

wb_name = '-Aggregated.xlsx'

cwd_dir = os.getcwd()
data_dir = os.getcwd()+"\\data\\"

death = 0
infected = 0
line = 1
line_total = 1

workbook = Workbook()
ws = workbook.active

sheet_wy     = workbook.create_sheet('WY', 0)
sheet_wi     = workbook.create_sheet('WI', 0)
sheet_wv     = workbook.create_sheet('WV', 0)
sheet_wa     = workbook.create_sheet('WA', 0)
sheet_va     = workbook.create_sheet('VA', 0)
sheet_vt     = workbook.create_sheet('VT', 0)
sheet_ut     = workbook.create_sheet('UT', 0)
sheet_tx     = workbook.create_sheet('TX', 0)
sheet_tn     = workbook.create_sheet('TN', 0)
sheet_sd     = workbook.create_sheet('SD', 0)
sheet_sc     = workbook.create_sheet('SC', 0)
sheet_ri     = workbook.create_sheet('RI', 0)
sheet_pr     = workbook.create_sheet('PR', 0)
sheet_pa     = workbook.create_sheet('PA', 0)
sheet_or     = workbook.create_sheet('OR', 0)
sheet_ok     = workbook.create_sheet('OK', 0)
sheet_oh     = workbook.create_sheet('OH', 0)
sheet_nc     = workbook.create_sheet('NC', 0)
sheet_ny     = workbook.create_sheet('NY', 0)
sheet_nm     = workbook.create_sheet('NM', 0)
sheet_nj     = workbook.create_sheet('NJ', 0)
sheet_nh     = workbook.create_sheet('NH', 0)
sheet_nv     = workbook.create_sheet('NV', 0)
sheet_ne     = workbook.create_sheet('NE', 0)
sheet_mt     = workbook.create_sheet('MT', 0)
sheet_mo     = workbook.create_sheet('MO', 0)
sheet_ms     = workbook.create_sheet('MS', 0)
sheet_mn     = workbook.create_sheet('MN', 0)
sheet_mi     = workbook.create_sheet('MI', 0)
sheet_ma     = workbook.create_sheet('MA', 0)
sheet_md     = workbook.create_sheet('MD', 0)
sheet_me     = workbook.create_sheet('ME', 0)
sheet_la     = workbook.create_sheet('LA', 0)
sheet_ky     = workbook.create_sheet('KY', 0)
sheet_ks     = workbook.create_sheet('KS', 0)
sheet_ia     = workbook.create_sheet('IA', 0)
sheet_in     = workbook.create_sheet('IN', 0)
sheet_il     = workbook.create_sheet('IL', 0)
sheet_id     = workbook.create_sheet('ID', 0)
sheet_hi     = workbook.create_sheet('HI', 0)
sheet_ga     = workbook.create_sheet('GA', 0)
sheet_fl     = workbook.create_sheet('FL', 0)
sheet_dc     = workbook.create_sheet('DC', 0)
sheet_de     = workbook.create_sheet('DE', 0)
sheet_ct     = workbook.create_sheet('CT', 0)
sheet_co     = workbook.create_sheet('CO', 0)
sheet_ca     = workbook.create_sheet('CA', 0)
sheet_ar     = workbook.create_sheet('AR', 0)
sheet_az     = workbook.create_sheet('AZ', 0)
sheet_as     = workbook.create_sheet('AS', 0)
sheet_ak     = workbook.create_sheet('AK', 0)
sheet_al     = workbook.create_sheet('AL', 0)
sheet_us     = workbook.create_sheet('US', 0)
sheet_totals = workbook.create_sheet('Projections', 0)

date_object = datetime.date.today()

sheet_totals['A'+str(line_total)] = 'US Death Count Projections'
line_total += 1

def analysis_state(sheet_name, state_name):
    global line, line_total
    print("Parsing State: "+state_name)
    death = 0
    infected = 0
    line = 1
    sheet_name['A'+str(line)] = (state_name)
    line += 1
    sheet_name['A'+str(line)] = ('Date')
    sheet_name['B'+str(line)] = ('Infected')
    sheet_name['C'+str(line)] = ('Rate')
    sheet_name['D'+str(line)] = ('Deaths')
    sheet_name['E'+str(line)] = ('Rate')
    sheet_name['F'+str(line)] = ('Avaerage (7-Day)')
    sheet_name.column_dimensions['A'].width = 10
    sheet_name.column_dimensions['F'].width = 16
    line += 1 
    for fname in os.listdir(data_dir):
        death = 0
        infected = 0
        recovered = 0
        average_7 = 0
        if fname.endswith(".csv"):
            date = fname.split('.')[0]
            with open(data_dir + fname) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                line_count = 0
                for row in csv_reader:
                    if row[3] == 'US':
                        if state_name == 'all':
                            death = death + int(row[8])
                            infected = infected + int(row[7])
                            recovered = recovered + int(row[9])
                        elif row[2].lower() == state_name:
                            death = death + int(row[8])
                            infected = infected + int(row[7])
                            recovered = recovered + int(row[9])
                    if row[1].lower() == 'us':
                        if state_name == 'all':
                            if row[4] != '':
                                death = death + int(row[4])
                            if row[3] != '':
                                infected = infected + int(row[3])
                            if row[5] != '':
                                recovered = recovered + int(row[5])
                        elif row[0].lower() == state_name:
                            if row[4] != '':
                                death = death + int(row[4])
                            if row[3] != '':
                                infected = infected + int(row[3])
                            if row[5] != '':
                                recovered = recovered + int(row[5])
                    line_count += 1
            sheet_name['A'+str(line)] = date
            sheet_name['B'+str(line)] = infected
            sheet_name['D'+str(line)] = death
            if sheet_name['B' + str(line-1)].value == 0:
                sheet_name['C' + str(line)] = '0'
                sheet_name['C'+str(line)].number_format = '00.00%'
            elif line == 3:
                sheet_name['C' + str(line)] = '0'
                sheet_name['C'+str(line)].number_format = '00.00%'
            else:
                sheet_name['C'+str(line)] = ('=(B' + str(line) + '/B' + str(line-1) + ') - 1')
                sheet_name['C'+str(line)].number_format = '00.00%'
            
            if sheet_name['D' + str(line-1)].value == 0:
                sheet_name['E' + str(line)] = '0'
                sheet_name['C'+str(line)].number_format = '00.00%'
            elif line == 3:
                sheet_name['E' + str(line)] = '0'
                sheet_name['C'+str(line)].number_format = '00.00%'
            else:
                sheet_name['E'+str(line)] = ('=(D' + str(line) + '/D' + str(line-1) + ') - 1')
                sheet_name['E'+str(line)].number_format = '00.00%'
            if sheet_name['E'+str(line)].value != '0':
                sheet_name['F'+str(line)] = ('=AVERAGE(E' + str(line-6) + ':E' + str(line)+')')
                average_7 = sheet_name['F'+str(line)].value
                sheet_name['F'+str(line)].number_format = '00.00%'
            else:
                sheet_name['F'+str(line)] = 0
                sheet_name['F'+str(line)].number_format = '00.00%'
            print ('Date: ' + date + ' Infected: ' + str(infected).rjust(7)
                + ' Death: ' + str(death).rjust(5)
                )
            line += 1
    line += 1
    sheet_name['A'+str(line)] = 'Average (7-Day):'
    sheet_name['C'+str(line)] = ('=AVERAGE(C' + str(line-8) + ':C' + str(line-2)+')')
    sheet_name['C'+str(line)].number_format = '00.00%'
    
    sheet_name['E'+str(line)] = ('=AVERAGE(E' + str(line-8) + ':E' + str(line-2)+')')
    sheet_name['E'+str(line)].number_format = '00.00%'

    #Projections
    line += 2
    sheet_name['A'+str(line)] = ('Projections')
    line += 1
    if (((sheet_name['D'+str(line-5)].value)  != 0) and 
        ((sheet_name['D'+str(line-6)].value)  != 0) and 
        ((sheet_name['D'+str(line-7)].value)  != 0) and 
        ((sheet_name['D'+str(line-8)].value)  != 0) and 
        ((sheet_name['D'+str(line-9)].value)  != 0) and 
        ((sheet_name['D'+str(line-10)].value) != 0) and 
        ((sheet_name['D'+str(line-11)].value) != 0)):
        average_7 = ((int(sheet_name['D'+str(line-5)].value)/ int(sheet_name['D'+str(line-6)].value)) + 
                    (int(sheet_name['D'+str(line-6)].value)/ int(sheet_name['D'+str(line-7)].value)) + 
                    (int(sheet_name['D'+str(line-7)].value)/ int(sheet_name['D'+str(line-8)].value)) + 
                    (int(sheet_name['D'+str(line-8)].value)/ int(sheet_name['D'+str(line-9)].value)) + 
                    (int(sheet_name['D'+str(line-9)].value)/ int(sheet_name['D'+str(line-10)].value)) + 
                    (int(sheet_name['D'+str(line-10)].value)/int(sheet_name['D'+str(line-11)].value)) + 
                    (int(sheet_name['D'+str(line-11)].value)/int(sheet_name['D'+str(line-12)].value))) / 7
    

    print()
    print()
    starting_death = int(sheet_name['D'+str(line-5)].value)
    starting_date = sheet_name['A'+str(line-5)].value

    print('Final Death Count (US): ' + str(starting_death))
    print('Seven Day Average Increase (US): ' + str(average_7))
    sheet_name['A'+str(line)] = ('Day +1')
    sheet_name['D'+str(line)] = (starting_death * average_7)
    sheet_name['E'+str(line)] = (((int(sheet_name['D'+str(line)].value)) / starting_death)-1)
    sheet_name['D'+str(line)].number_format = '0'
    sheet_name['E'+str(line)].number_format = '00.00%'
    line += 1
    sheet_name['A'+str(line)] = ('Day +2')
    sheet_name['D'+str(line)] = (sheet_name['D'+str(line-1)].value * average_7)
    sheet_name['E'+str(line)] = (((int(sheet_name['D'+str(line)].value)) / starting_death)-1)
    sheet_name['D'+str(line)].number_format = '0'
    sheet_name['E'+str(line)].number_format = '00.00%'
    line += 1
    sheet_name['A'+str(line)] = ('Day +3')
    sheet_name['D'+str(line)] = (sheet_name['D'+str(line-1)].value * average_7)
    sheet_name['E'+str(line)] = (((int(sheet_name['D'+str(line)].value)) / starting_death)-1)
    sheet_name['D'+str(line)].number_format = '0'
    sheet_name['E'+str(line)].number_format = '00.00%'
    line += 1
    sheet_name['A'+str(line)] = ('Day +4')
    sheet_name['D'+str(line)] = (sheet_name['D'+str(line-1)].value * average_7)
    sheet_name['E'+str(line)] = (((int(sheet_name['D'+str(line)].value)) / starting_death)-1)
    sheet_name['D'+str(line)].number_format = '0'
    sheet_name['E'+str(line)].number_format = '00.00%'
    line += 1
    sheet_name['A'+str(line)] = ('Day +5')
    sheet_name['D'+str(line)] = (sheet_name['D'+str(line-1)].value * average_7)
    sheet_name['E'+str(line)] = (((int(sheet_name['D'+str(line)].value)) / starting_death)-1)
    sheet_name['D'+str(line)].number_format = '0'
    sheet_name['E'+str(line)].number_format = '00.00%'
    line += 1
    sheet_name['A'+str(line)] = ('Day +6')
    sheet_name['D'+str(line)] = (sheet_name['D'+str(line-1)].value * average_7)
    sheet_name['E'+str(line)] = (((int(sheet_name['D'+str(line)].value)) / starting_death)-1)
    sheet_name['D'+str(line)].number_format = '0'
    sheet_name['E'+str(line)].number_format = '00.00%'
    line += 1
    sheet_name['A'+str(line)] = ('Day +7')
    sheet_name['D'+str(line)] = (sheet_name['D'+str(line-1)].value * average_7)
    sheet_name['E'+str(line)] = (((int(sheet_name['D'+str(line)].value)) / starting_death)-1)
    sheet_name['D'+str(line)].number_format = '0'
    sheet_name['E'+str(line)].number_format = '00.00%'

    #Add to totals sheet
    sheet_totals['A'+str(line_total)] = state_name.title()
    sheet_totals['B'+str(line_total)] = (average_7-1)
    sheet_totals['B'+str(line_total)].number_format = '00.00%'

    line_total += 1
    starting_death
    sheet_totals['A'+str(line_total)]   = ("Today's Count")
    sheet_totals['A'+str(line_total+1)] = (starting_death)

    sheet_totals['B'+str(line_total)]   = ('Day +1')
    sheet_totals['B'+str(line_total+1)] = (starting_death * average_7)
    sheet_totals['B'+str(line_total+2)] = (((int(sheet_totals['B'+str(line_total+1)].value)) / starting_death)-1)
    sheet_totals['B'+str(line_total+1)].number_format = '0'
    sheet_totals['B'+str(line_total+2)].number_format = '00.00%'
    
    sheet_totals['C'+str(line_total)]   = ('Day +2')

    sheet_totals['C'+str(line_total+1)] = (sheet_totals['B'+str(line_total+1)].value * average_7)
    sheet_totals['C'+str(line_total+2)] = (((int(sheet_totals['C'+str(line_total+1)].value)) / starting_death)-1)
    sheet_totals['C'+str(line_total+1)].number_format = '0'
    sheet_totals['C'+str(line_total+2)].number_format = '00.00%'
    
    sheet_totals['D'+str(line_total)]   = ('Day +3')
    sheet_totals['D'+str(line_total+1)] = (sheet_totals['C'+str(line_total+1)].value * average_7)
    sheet_totals['D'+str(line_total+2)] = (((int(sheet_totals['D'+str(line_total+1)].value)) / starting_death)-1)
    sheet_totals['D'+str(line_total+1)].number_format = '0'
    sheet_totals['D'+str(line_total+2)].number_format = '00.00%'
    
    sheet_totals['E'+str(line_total)]   = ('Day +4')
    sheet_totals['E'+str(line_total+1)] = (sheet_totals['D'+str(line_total+1)].value * average_7)
    sheet_totals['E'+str(line_total+2)] = (((int(sheet_totals['E'+str(line_total+1)].value)) / starting_death)-1)
    sheet_totals['E'+str(line_total+1)].number_format = '0'
    sheet_totals['E'+str(line_total+2)].number_format = '00.00%'
    
    sheet_totals['F'+str(line_total)]   = ('Day +5')
    sheet_totals['F'+str(line_total+1)] = (sheet_totals['E'+str(line_total+1)].value * average_7)
    sheet_totals['F'+str(line_total+2)] = (((int(sheet_totals['F'+str(line_total+1)].value)) / starting_death)-1)
    sheet_totals['F'+str(line_total+1)].number_format = '0'
    sheet_totals['F'+str(line_total+2)].number_format = '00.00%'
    
    sheet_totals['G'+str(line_total)]   = ('Day +6')
    sheet_totals['G'+str(line_total+1)] = (sheet_totals['F'+str(line_total+1)].value * average_7)
    sheet_totals['G'+str(line_total+2)] = (((int(sheet_totals['G'+str(line_total+1)].value)) / starting_death)-1)
    sheet_totals['G'+str(line_total+1)].number_format = '0'
    sheet_totals['G'+str(line_total+2)].number_format = '00.00%'

    sheet_totals['H'+str(line_total)]   = ('Day +7')
    sheet_totals['H'+str(line_total+1)] = (sheet_totals['G'+str(line_total+1)].value * average_7)
    sheet_totals['H'+str(line_total+2)] = (((int(sheet_totals['H'+str(line_total+1)].value)) / starting_death)-1)
    sheet_totals['H'+str(line_total+1)].number_format = '0'
    sheet_totals['H'+str(line_total+2)].number_format = '00.00%'

    line_total += 4
    
    print()
 
def collect_totals():
    list_sheets = workbook.sheetnames
    for i in range(len(list_sheets)):
        print(list_sheets[i])

analysis_state(sheet_us, 'all')
analysis_state(sheet_ak, 'alaska')
analysis_state(sheet_al, 'alabama')
analysis_state(sheet_ar, 'arkansas')
#analysis_state(sheet_as, 'american somoa')
analysis_state(sheet_az, 'arizona')
analysis_state(sheet_ca, 'california')
analysis_state(sheet_co, 'colorado')
analysis_state(sheet_ct, 'connecticut')
analysis_state(sheet_dc, 'district of columbia')
analysis_state(sheet_de, 'delaware')
analysis_state(sheet_fl, 'florida')
analysis_state(sheet_ga, 'georgia')
#analysis_state(sheet_hi, 'hawaii')
analysis_state(sheet_ia, 'iowa')
analysis_state(sheet_id, 'idaho')
analysis_state(sheet_il, 'illinois')
analysis_state(sheet_in, 'indiana')
analysis_state(sheet_ks, 'kansas')
analysis_state(sheet_ky, 'kentucky')
analysis_state(sheet_la, 'louisiana')
analysis_state(sheet_ma, 'massachusetts')
analysis_state(sheet_md, 'maryland')
analysis_state(sheet_me, 'maine')
analysis_state(sheet_mi, 'michigan')
analysis_state(sheet_mn, 'minnesota')
analysis_state(sheet_mo, 'missouri')
analysis_state(sheet_ms, 'mississippi')
analysis_state(sheet_mt, 'montana')
analysis_state(sheet_nc, 'north carolina')
analysis_state(sheet_ne, 'nebraska')
#analysis_state(sheet_nh, 'new hampshire')
analysis_state(sheet_nj, 'new jersey')
analysis_state(sheet_nm, 'new mexico')
analysis_state(sheet_nv, 'nevada')
analysis_state(sheet_ny, 'new york')
analysis_state(sheet_oh, 'ohio')
analysis_state(sheet_ok, 'oklahoma')
analysis_state(sheet_or, 'oregon')
analysis_state(sheet_pa, 'pennsylvania')
analysis_state(sheet_pr, 'puerto rico')
analysis_state(sheet_ri, 'rhode island')
analysis_state(sheet_sc, 'south carolina')
analysis_state(sheet_sd, 'south dakota')
analysis_state(sheet_tn, 'tennessee')
analysis_state(sheet_tx, 'texas')
analysis_state(sheet_ut, 'utah')
analysis_state(sheet_va, 'virginia')
analysis_state(sheet_vt, 'vermont')
analysis_state(sheet_wa, 'washington')
analysis_state(sheet_wi, 'wisconsin')
#analysis_state(sheet_wv, 'west virginia')
#analysis_state(sheet_wy, 'wyoming')


sheet_totals.column_dimensions['A'].width = 13
sheet_totals.column_dimensions['B'].width = 10
sheet_totals.column_dimensions['C'].width = 10
#sheet_totals.column_dimensions['B'].alignment = Alignment(horizontal='right')
#sheet_totals.column_dimensions['C'].alignment = Alignment(horizontal='right')

workbook.save(filename=cwd_dir + '\\' + str(date_object) + wb_name)

print()
print("Complete")
print()
print()


